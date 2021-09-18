from openpyxl import Workbook,load_workbook 
wb=load_workbook("marketing_campaign.xlsx")
ws = wb.active
import matplotlib.pyplot as plt
def main():
    row_letters=''
    number_of_PhD=0
    number_of_Graduation=0
    number_of_Basic=0
    number_of_Master=0
    for row in range (2,2242):
        row_string=str(ws.cell(row,3).value)
        if row_string=='PhD':
            number_of_PhD+=1
        elif row_string=='Basic':
            number_of_Basic+=1
        elif row_string=='Master':
            number_of_Master+=1
        else:
            number_of_Graduation+=1
    print('number of PhD:',number_of_PhD)
    print('number of Basic:',number_of_Basic)
    print('number of Master:',number_of_Master)
    print('number of Graduation:',number_of_Graduation)
    heights=[number_of_PhD,number_of_Basic,number_of_Master,number_of_Graduation]
    left_edges=[0,10,20,30]
    
    bar_width=9
    plt.bar(left_edges,heights,bar_width)
    plt.title('Education Levels of Workers')
    plt.xticks([0,10,20,30],
               ['PhD','Basic','Master','Graduation'])
    plt.show()
    
main()
    
    
    
            